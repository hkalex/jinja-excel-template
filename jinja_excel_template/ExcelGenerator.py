from lxml import etree

from openpyxl import Workbook
import json

from importlib import import_module
from jinja_excel_template.helper.setSheetAttr import set_sheet_attr
from jinja_excel_template.helper.setCellStyle import set_cell_style
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from jinja_excel_template.helper.MergeCell import merge_cell

from jinja_excel_template.helper.DimensionHelper import set_column_row_dimensions



class ExcelGenerator:
    """
    This class generates Excel file from ExcelXML
    """

    options = {}

    def __find_class(self, class_str:str):
        # class_str: str = 'A.B.YourClass'
        try:
            module_path, class_name = class_str.rsplit('.', 1)
            module = import_module(module_path)
            return getattr(module, class_name)
        except (ImportError, AttributeError) as e:
            raise ImportError(class_str)


    def generate_excel(self, output_file_path, xml_string = None, xml_file_path = None):
        if xml_string != None:
            self.xml_Excel = etree.fromstring(xml_string)
        elif xml_file_path != None:
            f = open(xml_file_path, 'r')
            self.xml_Excel = etree.fromstring(f.read())
        else:
            raise Exception("xml_string and xml_file_path cannot be both None")
        

        workbook = Workbook()
        default_sheet = workbook.active
        
        for sheet_index, xml_sheet in enumerate(self.xml_Excel.xpath("sheet")):
            sheet_name = xml_sheet.get("sheet_name")
            sheet_type = xml_sheet.get("type")
            if not sheet_type:
                worksheet = workbook.create_sheet(sheet_name, sheet_index)
            elif sheet_type == "chart_sheet":
                worksheet = workbook.create_chartsheet(sheet_name, sheet_index)
                

            # set worksheet attributes
            set_sheet_attr(worksheet, xml_sheet)

            for row_index, xml_row in enumerate(xml_sheet.xpath("row")):
                for cell_index, xml_cell in enumerate(xml_row.xpath("cell")):

                    cell_type_name = xml_cell.get("type") or "General"

                    # This only loads the class in the cell_type folder
                    # This can be future extendable
                    cell_type_class = self.__find_class('jinja_excel_template.cell_type.' + cell_type_name + '.' + cell_type_name)

                    cell_type_handler = cell_type_class()
                    cell_type_handler.set_cell(worksheet, row_index+1, cell_index+1, xml_cell)

                    # set styles
                    if type(worksheet) == Worksheet:
                        set_cell_style(worksheet.cell(row=row_index+1, column=cell_index+1), xml_cell)

                    # merge cell
                    merge_cell(worksheet, row_index+1, cell_index+1, xml_cell)

            
            # set set_column_row_dimensions
            set_column_row_dimensions(worksheet, xml_sheet)



        # remove the default sheet
        workbook.remove(default_sheet)
        workbook.save(output_file_path)

        return