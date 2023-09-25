from openpyxl.worksheet.worksheet import Worksheet
from jinja_excel_template.helper.GetFont import get_font

from jinja_excel_template.helper.utils import is_float


class Number:
    def set_cell(self, worksheet:Worksheet, row:int, column:int, xml_cell):
        """
        This cell type set the cell as general cell. row and column are 1 based index.
        """
        if xml_cell != None:
            str_value = xml_cell.get("value")

        if is_float(str_value):
            float_value = float(str_value)
            if float_value.is_integer():
                worksheet.cell(row, column, int(str_value))
            else:
                worksheet.cell(row, column, float_value)
        else:
            worksheet.cell(row, column, str_value)

