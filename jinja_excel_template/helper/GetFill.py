from openpyxl.styles import PatternFill
import inspect

attr_dict = {
    "fill_type": str,
    "start_color": str,
    "end_color": str
}

def get_fill(xml_cell):

# >>> fill = PatternFill(fill_type=None,
# ...                 start_color='FFFFFFFF',
# ...                 end_color='FF000000')

    fill = PatternFill()

    # for (name, value) in inspect.getmembers(fill):
    #     print(name + "=" + str(type(value)))

    # return

    prefix = "fill."
    for attr in attr_dict:
        if xml_cell.get(prefix + attr) != None:
            setattr(fill, attr, attr_dict.get(attr)(xml_cell.get(prefix + attr)))

    return fill
