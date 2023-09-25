from openpyxl.worksheet.worksheet import Worksheet
from lxml.etree import _Element
from openpyxl.utils import get_column_letter
from jinja_excel_template.helper.typeconverter import str2bool

def set_column_row_dimensions(ws:Worksheet, xml_sheet:_Element):
    """
    This function sets the column width, column hidden, row height and row hidden.
    """
    _set_column_width(ws, xml_sheet)
    _set_row_height(ws, xml_sheet)


def _set_column_width(ws:Worksheet, xml_sheet:_Element):
    for column_index, xml_column in enumerate(xml_sheet.xpath("columns/column")):
        width = xml_column.get("width")
        if width != None and str.isnumeric(width):
            ws.column_dimensions[get_column_letter(column_index+1)].width = float(width)
        elif width != None and width.lower() == "auto":
            max_len = 0
            for row in ws.rows:
                cell = row[column_index]
                if cell.value:
                    max_len = max( max_len, len(str(cell.value)) )
            
            ws.column_dimensions[get_column_letter(column_index+1)].width = float(max_len)
        
        hidden = xml_column.get('hidden')
        if hidden != None:
            ws.column_dimensions[get_column_letter(column_index+1)].hidden = str2bool(hidden)



def _set_row_height(ws:Worksheet, xml_sheet:_Element):
    for row_index, xml_row in enumerate(xml_sheet.xpath("row")):
        height = xml_row.get('height')
        if height != None and str.isnumeric(height):
            ws.row_dimensions[row_index+1].height = float(height)
        hidden = xml_row.get('hidden')
        if hidden != None:
            ws.row_dimensions[row_index+1].hidden = str2bool(hidden)